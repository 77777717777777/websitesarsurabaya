"""
Parsing file Excel laporan kejadian SAR (format asli laporan bulanan/tahunan)
menjadi baris-baris siap divalidasi oleh normalize_payload() di
routes/admin_routes.py.

Logika di file ini adalah PORTING LANGSUNG dari script ETL Python yang sudah
ditulis, dipakai, dan diuji sendiri oleh pengguna (notebook etl.ipynb) untuk
mengubah laporan mentah (satu sheet per bulan, berisi baris kategori + baris
bernomor per-kejadian + baris lanjutan teks) menjadi data flat kejadian_sar.
Fungsi-fungsi di bawah sengaja dibuat semirip mungkin dengan versi aslinya
(nama fungsi & urutan logika dipertahankan) supaya mudah dicocokkan/diaudit
terhadap notebook sumbernya kalau ada perbedaan hasil.

Sumber format mentah TIDAK punya header kolom sederhana per field -- setiap
sheet adalah satu bulan, dengan baris kategori ("KECELAKAAN KAPAL", dst),
baris kejadian bernomor (1., 2., dst di kolom A), dan kadang baris lanjutan
teks tanpa nomor (kolom instansi/personel/lokasi ditemukan yang isinya
panjang, disambung ke baris di bawahnya). Makanya parsing dilakukan baris
demi baris dengan aturan yang sama seperti script aslinya, BUKAN pencocokan
nama header seperti pendekatan awal (client-side, sudah digantikan file ini).
"""
import re
from datetime import datetime, timedelta

import openpyxl
from geopy.distance import distance
from geopy.point import Point


# ============================================================
# 1. Konstanta (persis dari script ETL asli pengguna)
# ============================================================

MONTHS = [
    "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
    "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER",
]

CATEGORY_KEYWORDS = [
    "KECELAKAAN PESAWAT UDARA",
    "KECELAKAAN KAPAL",
    "BENCANA",
    "KONDISI YANG MEMBAHAYAKAN JIWA MANUSIA",
    "KECELAKAAN DGN PENANGANAN KHUSUS",
    "PENGECEKAN SIGNAL DISTRESS",
]

STOP_MARKERS = ["response time"]

BULAN_DICT = {
    "JANUARI": 1, "JAN": 1, "FEBRUARI": 2, "FEB": 2, "MARET": 3, "MAR": 3,
    "APRIL": 4, "APR": 4, "MEI": 5, "MAY": 5, "JUNI": 6, "JUN": 6,
    "JULI": 7, "JUL": 7, "AGUSTUS": 8, "AGU": 8, "AUG": 8,
    "SEPTEMBER": 9, "SEP": 9, "OKTOBER": 10, "OKT": 10, "OCT": 10,
    "NOVEMBER": 11, "NOV": 11, "DESEMBER": 12, "DES": 12, "DEC": 12,
}

BULAN_NAMA = [
    'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember',
]

# Titik acuan default kalau koordinat ditulis dalam format Radial (arah +
# jarak dari markas), dipakai kalau tidak bisa menentukan pos asal lain.
POS_SAR_SURABAYA = Point(-7.3846, 112.7725)

_KAPAL_SURABAYA = [r'permadi', r'widura', r'antasena', r'kahyangan']

WILAYAH_MAP = {
    'Surabaya': [r'surabaya'] + _KAPAL_SURABAYA,
    'Sumenep': [r'sumenep'],
    'Trenggalek': [r'trenggalek'],
    'Jember': [r'jember'],
    'Banyuwangi': [r'banyuwangi'],
    'Malang': [r'malang'],
    'Bojonegoro': [r'bojonegoro', r'bojongoro'],
    'Lamongan': [r'lamongan'],
}

KATEGORI_MAP = [
    ('Kecelakaan Lalu Lintas',        [r'ranmor', r'laka.?lantas', r'kecelakaan lalu lintas', r'lala lantas']),
    ('Bencana Alam - Banjir',         [r'banjir']),
    ('Bencana Alam - Longsor',        [r'longsor']),
    ('Bencana Alam - Gempa',          [r'gempa']),
    ('Percobaan Bunuh Diri',          [r'bunuh diri', r'menceburkan diri']),
    ('Pesawat Jatuh',                 [r'pesawat jatuh']),
    ('Kapal - Terbakar',              [r'kapal.*terbakar', r'terbakar.*kapal', r'^km .*terbakar']),
    ('Kapal - Tenggelam/Terbalik/Kandas/Rusak', [
        r'kapal.*tenggelam', r'tenggelam.*kapal', r'kapal.*terbalik', r'terbalik.*kapal',
        r'kapal.*kandas', r'kapal.*terdampar', r'kapal.*bocor', r'kapal.*pecah', r'kapal.*miring',
        r'kapal.*mati mesin', r'kpl.*mati mesin', r'kapal.*tubrukan', r'kapal.*bertubrukan',
        r'perahu.*terbalik', r'perahu.*tenggelam', r'perahu.*bocor', r'kapal/ ?perahu.*tenggelam',
        r'perahu.*dihantam ombak', r'kecelakaan kapal',
    ]),
    ('Kapal - Hilang Kontak',         [r'kapal.*(hilang kontak|lo[cs]t contact)', r'(hilang kontak|lo[cs]t contact).*kapal',
                                        r'^lo[cs]t contact\.?$', r'^hilang kontak$']),
    ('Man Over Boat (MOB)',           [r'man over boat', r'^mob$']),
    ('Orang Tenggelam/Hanyut/Tercebur', [
        r'tengg?elam', r'hanyut', r'tercebur', r'terjebur', r'terjebak.*sungai',
    ]),
    ('Orang Terseret Arus/Ombak',     [r'ters[e]?ret', r'teseret', r'tersapu ombak', r'terbawa arus']),
    ('Orang Hilang/Tersesat',         [r'hilang', r'tersesat']),
    ('Orang Terjatuh',                [r'terjatuh', r'jatuh', r'terjun', r'terpeleset']),
    ('Orang Tertimpa/Terjepit/Terjebak', [r'tertimpa', r'terjepit', r'terjebak di ruang']),
    ('Evakuasi/Medevac',              [r'evakuasi', r'medevac']),
]


# ============================================================
# 2. Helper dasar baca sel & teks (persis dari script asli)
# ============================================================

def cellval(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def is_category_row(a_val):
    if a_val is None:
        return None
    text = str(a_val).strip().upper()
    for kw in CATEGORY_KEYWORDS:
        if text == kw or text.startswith(kw):
            return str(a_val).strip()
    return None


def is_incident_number(a_val):
    if a_val is None:
        return False
    if isinstance(a_val, (int, float)):
        return True
    s = str(a_val).strip()
    if s in ("-", ""):
        return False
    return bool(re.match(r"^\d+\.?\d*\.?$", s))


def append_text(existing, new):
    if new is None:
        return existing
    if existing is None:
        return str(new)
    return str(existing) + "\n" + str(new)


def fmt_num(v):
    if v is None:
        return None
    if isinstance(v, float):
        if v == int(v):
            return int(v)
        return v
    return v


def num_or_none(v):
    """Untuk kolom numerik (POB, S/MD/H) -- sumber sering menulis '-' atau
    'nihil' untuk nol/kosong (lihat is_missing()), bukan angka. Itu HARUS
    jadi None di sini, bukan diteruskan sebagai teks '-', supaya validasi
    angka di normalize_payload() tidak salah menganggapnya error."""
    if is_missing(v):
        return None
    return fmt_num(v)


def safe_float(v):
    """Ekstraksi angka (mis. dari kolom jarak) dengan aman."""
    if v is None:
        return 0.0
    try:
        clean_val = re.sub(r'[^\d.]', '', str(v))
        return float(clean_val) if clean_val else 0.0
    except ValueError:
        return 0.0


def combine_val_unit(val, unit):
    val_s = None if val is None else (str(int(val)) if isinstance(val, float) and val == int(val) else str(val))
    unit_s = str(unit).strip() if unit is not None else None
    if val_s is None and unit_s is None:
        return None
    if val_s is None:
        return unit_s
    if unit_s is None:
        return val_s
    return f"{val_s} {unit_s}"


def parse_instansi(raw_t, raw_u):
    if not raw_t:
        return None
    raw_u_s = "" if raw_u is None else str(raw_u)
    counts = [c.strip() for c in raw_u_s.split("\n") if c.strip()]
    lines_t = str(raw_t).split("\n")

    numbered_lines = sum(1 for line in lines_t if re.match(r"^(\d+\.|-)\s*", line.strip()))
    is_numbered_list = numbered_lines >= 2

    instansi_list = []
    current_name = ""

    for t in lines_t:
        t = t.strip()
        if not t:
            continue
        if is_numbered_list:
            if re.match(r"^(\d+\.|-)\s*", t):
                if current_name:
                    instansi_list.append(current_name)
                current_name = re.sub(r"^(\d+\.|-)\s*", "", t).strip(" ;.")
            else:
                if current_name:
                    current_name += " " + t.strip(" ;.")
                else:
                    current_name = t.strip(" ;.")
        else:
            instansi_list.append(t.strip(" ;."))

    if current_name and is_numbered_list:
        instansi_list.append(current_name)

    pairs = []
    for i, name in enumerate(instansi_list):
        cnt = counts[i] if i < len(counts) else ""
        if cnt and cnt != "-":
            pairs.append(f"{name}({cnt})")
        else:
            pairs.append(name)

    return ", ".join(pairs) if pairs else None


def is_missing(v):
    if v is None:
        return True
    s = str(v).strip()
    if s in ("", "-", "nihil"):
        return True
    return False


# ============================================================
# 3. Ekstraksi teks koordinat & konversi ke desimal
# ============================================================

def extract_koordinat(raw_text):
    if not raw_text:
        return None
    coord_pattern = r"\d{1,3}[^\da-zA-Z]{0,5}\d{1,2}[^a-zA-Z]{0,25}[SsUuNn](?![a-zA-Z])[^a-zA-Z\d]{0,10}\d{1,3}[^\da-zA-Z]{0,5}\d{1,2}[^a-zA-Z]{0,25}(?:[EeBbTtWw](?![a-zA-Z]))?"
    coord_match = re.search(coord_pattern, str(raw_text))
    if coord_match:
        return coord_match.group(0).strip()

    fallback_match = re.search(r"(?i)(?:koord|kord)[a-z]*[:\s]+([^\nA-Za-z]*(?:[SUN]\b)?[^\nA-Za-z]*(?:[EBTW]\b)?)", str(raw_text))
    if fallback_match and len(fallback_match.group(1).strip()) > 5:
        return fallback_match.group(1).strip()

    radial_match = re.search(r"Radial\s*[\d.,]+\s*°?", str(raw_text), re.IGNORECASE)
    if radial_match:
        return radial_match.group(0)
    return None


def extract_koordinat_ditemukan(raw_text):
    if not raw_text:
        return "-"
    text = str(raw_text).strip()
    coord = extract_koordinat(text)
    if coord:
        return coord

    negation_pattern = r"(?i)b[e]?l[u]?m\.?\s+(?:berhasil\s+|dapat\s+|bisa\s+)?(?:menemukan|ditemukan|diketemukan|mengevakuasi|dievakuasi|temukan)(?:\s+dan\s+(?:menemukan|ditemukan|diketemukan|mengevakuasi|dievakuasi|temukan))?"
    text_clean = re.sub(negation_pattern, "", text)

    ada_ditemukan = bool(re.search(r"(?i)(temukan|mukan|evakuasi|keadaan md|meninggal dunia|\bmd\b)", text_clean))
    ada_hilang = bool(re.search(r"(?i)(dinyatakan hilang|korban hilang|\bhilang\b)", text_clean))
    ada_selamat = bool(re.search(r"(?i)(korban selamat|tidak ada korban jiwa|\bselamat\b)", text_clean))

    if ada_ditemukan or (ada_hilang and ada_selamat):
        cut_rule = r"(?=\n|$|,\s|\.?\s*(?:selanjutnya|kemudian|dievakuasi|dibawa|lalu|pukul|tim\s+sar|unsur|jenazah)\b)"
        jarak_matches = re.findall(rf"(?i)jarak\s+(.*?){cut_rule}", text)
        if jarak_matches:
            unique_j = []
            for m in jarak_matches:
                c = re.sub(r"[,.\s]+$", "", m.strip())
                if c and c not in unique_j:
                    unique_j.append(c)
            if unique_j:
                return "\n".join(unique_j)

        sekitar_matches = re.findall(rf"(?i)sekitar\s+(.*?){cut_rule}", text)
        if sekitar_matches:
            unique_s = []
            for m in sekitar_matches:
                c = re.sub(r"[,.\s]+$", "", m.strip())
                if c and c not in unique_s:
                    unique_s.append(c)
            if unique_s:
                return "\n".join([f"sekitar {s}" for s in unique_s])
    return "-"


def hitung_koordinat_desimal(teks, jarak_km=0.0):
    """Konversi teks koordinat (DMS/DDM atau Radial) ke desimal. Anti-typo:
    kalau arah bujur (E/W) lupa ditulis operator, diasumsikan 'E' (Timur) --
    wilayah kerja SAR Surabaya seluruhnya di belahan timur."""
    if not teks or str(teks).strip() in ("-", ""):
        return {"tipe": "-", "lat": None, "lon": None}

    teks = str(teks).strip()

    def parse_float(val_str):
        if not val_str:
            return 0.0
        parts = str(val_str).split('.')
        if len(parts) > 2:
            val_str = parts[0] + '.' + ''.join(parts[1:])
        try:
            return float(val_str)
        except ValueError:
            return 0.0

    dms_pattern = r"(\d+)[^\dA-Za-z]+(\d+(?:\.\d+)?)[^\dA-Za-z]*(?:([\d.]+)[^\dA-Za-z]*)?([NS])[\s\-,/]*(\d+)[^\dA-Za-z]+(\d+(?:\.\d+)?)[^\dA-Za-z]*(?:([\d.]+)[^\dA-Za-z]*)?([EW])?"
    match_dms = re.search(dms_pattern, teks, re.IGNORECASE)

    if match_dms:
        lat_deg = match_dms.group(1)
        lat_min = match_dms.group(2)
        lat_sec = match_dms.group(3) if match_dms.group(3) else 0.0
        lat_dir = match_dms.group(4)

        lon_deg = match_dms.group(5)
        lon_min = match_dms.group(6)
        lon_sec = match_dms.group(7) if match_dms.group(7) else 0.0
        lon_dir = match_dms.group(8) if match_dms.group(8) else 'E'

        lat = parse_float(lat_deg) + (parse_float(lat_min) / 60) + (parse_float(lat_sec) / 3600)
        if lat_dir.upper() == 'S':
            lat *= -1

        lon = parse_float(lon_deg) + (parse_float(lon_min) / 60) + (parse_float(lon_sec) / 3600)
        if lon_dir.upper() == 'W':
            lon *= -1

        return {"tipe": "DMS", "lat": round(lat, 6), "lon": round(lon, 6)}

    radial_match = re.search(r"Radial\s*([\d.]+)\s*°?", teks, re.IGNORECASE)
    if radial_match:
        radial_deg = parse_float(radial_match.group(1))
        titik_awal = POS_SAR_SURABAYA

        if jarak_km > 0:
            titik_tujuan = distance(kilometers=jarak_km).destination(point=titik_awal, bearing=radial_deg)
            return {"tipe": "RADIAL", "lat": round(titik_tujuan.latitude, 6), "lon": round(titik_tujuan.longitude, 6)}
        return {"tipe": "RADIAL_NO_JARAK", "lat": None, "lon": None}

    return {"tipe": "TIDAK_DIKENAL", "lat": None, "lon": None}


# ============================================================
# 4. Standardisasi waktu (porting dari standarisasi_waktu, cell 1 notebook)
# ============================================================

def standarisasi_waktu(teks, tahun_default, bulan_sheet_angka=None):
    """Ubah teks waktu mentah ('1231 1400 G' / '1 Januari 2023 Pukul 14.00' /
    objek datetime asli) jadi string 'YYYY-MM-DD HH:MM:SS'.

    Tambahan dari versi notebook: kalau bulan hasil parse jauh lebih besar
    dari bulan sheet-nya (mis. sheet JANUARI tapi kode waktu bulan 12 / "1231"),
    itu berarti kejadian terjadi di penghujung TAHUN SEBELUMNYA dan baru
    dilaporkan/diproses di laporan bulan berikutnya -- tahun dikoreksi mundur
    satu. Ini konsisten dengan data historis yang sudah ada di database (lihat
    baris pertama data 2023: kejadian '1231 1400 G' di sheet JANUARI 2023
    tersimpan sebagai 2022-12-31, bukan 2023-12-31)."""
    if teks is None:
        return None
    if isinstance(teks, datetime):
        return teks.strftime('%Y-%m-%d %H:%M:%S')

    s = str(teks).strip()
    if s.upper() in ("-", "", "NAN", "NONE"):
        return None

    teks_u = s.upper()

    pola_teks = r"(\d{1,2})\s+([A-Z]+)(?:\s+(\d{4}))?[,\s]*(?:PUKUL|JAM)?\s*(\d{1,2})[.:](\d{2})"
    match_teks = re.search(pola_teks, teks_u)

    if match_teks:
        tgl = int(match_teks.group(1))
        bln_teks = match_teks.group(2)
        thn = int(match_teks.group(3)) if match_teks.group(3) else int(tahun_default)
        jam = int(match_teks.group(4))
        mnt = int(match_teks.group(5))

        bln = BULAN_DICT.get(bln_teks)
        if bln is None:
            return s
        try:
            return f"{thn:04d}-{bln:02d}-{tgl:02d} {jam:02d}:{mnt:02d}:00"
        except ValueError:
            return s

    # Pola sandi waktu "MMDD HHMM G" (mis. '1231 1400 G'). Dibuat sedikit
    # lebih longgar dari catatan aslinya di notebook -- data sungguhan
    # kadang salah ketik pemisahnya (mis. '0321`1540 G' pakai backtick,
    # bukan spasi) atau lupa nulis huruf 'G' di akhir (mis. '0903 2100').
    # Bagian 'G' & jumlah spasi dibuat opsional/fleksibel supaya kasus itu
    # tetap terbaca, tapi pola 4+4 digit-nya sendiri tetap wajib persis --
    # jadi tidak akan salah tangkap teks lain yang kebetulan ada angkanya.
    pola_sandi = r"(\d{2})(\d{2})[^0-9A-Z]{0,3}(\d{2})(\d{2})\s*G?"
    match_sandi = re.search(pola_sandi, teks_u)

    if match_sandi:
        bln = int(match_sandi.group(1))
        tgl = int(match_sandi.group(2))
        jam = int(match_sandi.group(3))
        mnt = int(match_sandi.group(4))

        if 1 <= bln <= 12 and 1 <= tgl <= 31 and 0 <= jam <= 23 and 0 <= mnt <= 59:
            thn = int(tahun_default)
            if bulan_sheet_angka:
                if bln - bulan_sheet_angka >= 2:
                    thn -= 1
                elif bulan_sheet_angka - bln >= 10:
                    thn += 1
            try:
                return f"{thn:04d}-{bln:02d}-{tgl:02d} {jam:02d}:{mnt:02d}:00"
            except ValueError:
                return s

    return s


def _to_dt(std_str):
    """Parse hasil standarisasi_waktu ('YYYY-MM-DD HH:MM:SS') jadi datetime,
    atau None kalau bukan format itu (mis. teks mentah yang gagal dikenali)."""
    if not std_str:
        return None
    try:
        return datetime.strptime(std_str, '%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return None


# ============================================================
# 5. Turunan lain (porting dari cell 44/48 notebook)
# ============================================================

def cari_wilayah(teks):
    teks_lower = str(teks or '').lower()
    hasil = []
    for wilayah, pola_list in WILAYAH_MAP.items():
        for pola in pola_list:
            if re.search(pola, teks_lower):
                hasil.append(wilayah)
                break
    return hasil


def klasifikasi_kejadian(teks):
    teks_lower = str(teks or '').lower()
    for label, pola_list in KATEGORI_MAP:
        for pola in pola_list:
            if re.search(pola, teks_lower):
                return label
    return None


def ekstrak_menit(teks):
    if teks is None:
        return None
    match = re.search(r'(\d+)', str(teks))
    if match:
        return int(match.group(1))
    return None


def hitung_rentang_menit(waktu_lapor_std, waktu_kejadian_std):
    """RENTANG_WAKTU: selisih menit WAKTU LAPOR - WAKTU KEJADIAN. None kalau
    salah satu tidak bisa diparse; None juga (bukan string error) kalau
    hasilnya negatif -- ditandai lewat elemen kedua tuple return."""
    dt_lapor = _to_dt(waktu_lapor_std)
    dt_kejadian = _to_dt(waktu_kejadian_std)
    if dt_lapor is None or dt_kejadian is None:
        return None, False
    delta = dt_lapor - dt_kejadian
    total_menit = int(delta.total_seconds() // 60)
    if total_menit < 0:
        return None, True  # 'Data tidak valid' pada script asli -- disimpan NULL, ditandai lewat flag
    return total_menit, False


# ============================================================
# 6. Parsing sheet bulanan (porting dari parse_month_sheet)
# ============================================================

def parse_month_sheet(ws, tahun):
    bulan = ws.title.strip().upper()
    records = []
    current_category = None

    max_row = ws.max_row
    for r in range(8, max_row + 1):
        a_val = cellval(ws, r, 1)
        a_text = str(a_val).strip().lower() if a_val is not None else ""

        if any(marker in a_text for marker in STOP_MARKERS):
            break

        cat = is_category_row(a_val)
        if cat:
            current_category = cat
            continue

        if is_incident_number(a_val):
            row_cells = {c: cellval(ws, r, c) for c in range(1, 31)}
            rec = {
                "sheet": ws.title,
                "source_row": r,
                "tahun": tahun,
                "bulan": bulan,
                "kategori": current_category,
                "klasifikasi": row_cells.get(2),
                "jenis_kecelakaan": row_cells.get(3),
                "posisi_raw": row_cells.get(4),
                "kejadian_wkt": row_cells.get(5),
                "lapor": row_cells.get(6),
                "berangkat": row_cells.get(7),
                "tiba": row_cells.get(8),
                "selesai": row_cells.get(9),
                "jarak_val": row_cells.get(10),
                "jarak_unit": row_cells.get(11),
                "siap_val": row_cells.get(12),
                "siap_unit": row_cells.get(13),
                "tempuh_val": row_cells.get(14),
                "tempuh_unit": row_cells.get(15),
                "pob": row_cells.get(16),
                "selamat": row_cells.get(17),
                "md": row_cells.get(18),
                "hilang": row_cells.get(19),
                "instansi": row_cells.get(20),
                "personel": row_cells.get(21),
                "peralatan": row_cells.get(22),
                "sumber_berita": row_cells.get(23),
                "kendala": row_cells.get(24),
                "lok_ditemukan": row_cells.get(25),
                "lain_lain": row_cells.get(26),
                "biaya": row_cells.get(27),
                "aksi": cellval(ws, r, 29),
                "durasi": cellval(ws, r, 30),
            }
            records.append(rec)
            continue

        if records:
            current_row = records[-1]
            t_cont = cellval(ws, r, 20)
            u_cont = cellval(ws, r, 21)
            y_cont = cellval(ws, r, 25)
            if t_cont is not None:
                current_row["instansi"] = append_text(current_row["instansi"], t_cont)
            if u_cont is not None:
                current_row["personel"] = append_text(current_row["personel"], u_cont)
            if y_cont is not None:
                current_row["lok_ditemukan"] = append_text(current_row["lok_ditemukan"], y_cont)

    return records


# ============================================================
# 7. Rakit satu record hasil parse_month_sheet -> payload kejadian_sar
#    (bentuk field ini SAMA PERSIS dengan yang diharapkan normalize_payload
#    di routes/admin_routes.py, termasuk kolom "legacy" teks yang juga
#    dipakai data historis di database).
# ============================================================

def build_payload_row(rec):
    tahun = rec["tahun"]
    bulan_angka = BULAN_DICT.get(rec["bulan"].strip().upper().split()[0]) if rec.get("bulan") else None

    instansi_personel = parse_instansi(rec["instansi"], rec["personel"])

    koordinat_ext = extract_koordinat(rec["posisi_raw"])
    koordinat_ditemukan = extract_koordinat_ditemukan(rec["lok_ditemukan"])

    jarak_km = safe_float(rec["jarak_val"])
    lkk_des = hitung_koordinat_desimal(koordinat_ext, jarak_km)
    temu_des = hitung_koordinat_desimal(koordinat_ditemukan, 0.0)

    waktu_kejadian_std = standarisasi_waktu(rec["kejadian_wkt"], tahun, bulan_angka)
    waktu_lapor_std = standarisasi_waktu(rec["lapor"], tahun, bulan_angka)
    waktu_berangkat_std = standarisasi_waktu(rec["berangkat"], tahun, bulan_angka)
    waktu_tiba_std = standarisasi_waktu(rec["tiba"], tahun, bulan_angka)
    waktu_selesai_std = standarisasi_waktu(rec["selesai"], tahun, bulan_angka)

    rentang_menit, _rentang_invalid = hitung_rentang_menit(waktu_lapor_std, waktu_kejadian_std)

    kategori_kejadian = klasifikasi_kejadian(rec["klasifikasi"])
    wilayah_hasil = cari_wilayah(rec["aksi"])
    wilayah_mapped = ", ".join(wilayah_hasil) if wilayah_hasil else None

    tempuh_combined = combine_val_unit(rec["tempuh_val"], rec["tempuh_unit"])
    waktu_tempuh_menit = ekstrak_menit(tempuh_combined) if tempuh_combined else (
        rec["tempuh_val"] if isinstance(rec["tempuh_val"], (int, float)) else None
    )

    # WAKTU SIAP: kolom ini di database berupa angka menit polos (bukan teks
    # "20 Menit") -- ambil nilai numeriknya saja, sesuai bentuk historis.
    waktu_siap = rec["siap_val"] if isinstance(rec["siap_val"], (int, float)) else ekstrak_menit(rec["siap_val"])

    # DURASI OPERASI: kalau sumber Excel sudah mengisi kolom ini (biasanya
    # dihitung manual di lapangan), pakai apa adanya -- lebih dipercaya
    # daripada hasil hitung ulang dari berangkat/tiba di normalize_payload().
    durasi_raw = rec["durasi"]
    durasi_operasi_hari = None
    if isinstance(durasi_raw, (int, float)):
        durasi_operasi_hari = float(durasi_raw)
    elif isinstance(durasi_raw, str) and durasi_raw.strip():
        m = re.search(r'[\d.,]+', durasi_raw)
        if m:
            try:
                durasi_operasi_hari = float(m.group(0).replace(',', '.'))
            except ValueError:
                durasi_operasi_hari = None

    payload = {
        # Periode laporan (sheet asal) -- BUKAN diturunkan dari waktu_kejadian.
        # Konsisten dengan data historis: kejadian yang terjadi di penghujung
        # tahun/bulan sebelumnya tapi baru diproses di laporan bulan berikutnya
        # (lihat rollover di standarisasi_waktu di atas) tetap dicatat di bawah
        # periode SHEET-nya, bukan tanggal kalender waktu_kejadian -- ini yang
        # dipakai oleh filter/agregat tahun & bulan di dashboard publik.
        'tahun_sheet': tahun,
        'bulan_sheet_angka': bulan_angka,
        # Info kejadian
        'waktu_kejadian': waktu_kejadian_std,
        'kategori': rec["kategori"],
        'klasifikasi': rec["klasifikasi"],
        'kategori_kejadian': kategori_kejadian,
        'jenis_kecelakaan': rec["jenis_kecelakaan"],
        'posisi_koordinat_area': rec["posisi_raw"],
        # Lokasi kejadian (LKK)
        'koordinat_lkk_teks': koordinat_ext,
        'tipe_lkk': lkk_des["tipe"],
        'latitude_lkk': lkk_des["lat"],
        'longitude_lkk': lkk_des["lon"],
        'wilayah_mapped': wilayah_mapped,
        'sumber_berita': rec["sumber_berita"],
        # Waktu pelaksanaan
        'waktu_lapor': waktu_lapor_std,
        'waktu_berangkat': waktu_berangkat_std,
        'waktu_tiba': waktu_tiba_std,
        'waktu_selesai': waktu_selesai_std,
        'rentang_waktu': rentang_menit,
        'jarak': combine_val_unit(rec["jarak_val"], rec["jarak_unit"]),
        'waktu_siap': waktu_siap,
        'waktu_tempuh': tempuh_combined,
        'waktu_tempuh_menit': waktu_tempuh_menit,
        # Korban
        'pob': num_or_none(rec["pob"]),
        's_org': num_or_none(rec["selamat"]),
        'md_org': num_or_none(rec["md"]),
        'h_org': num_or_none(rec["hilang"]),
        # Lain-lain
        'instansi_jml_person': instansi_personel,
        'peralatan': rec["peralatan"],
        'kendala_pelaksanaan_ops_sar': rec["kendala"],
        'lokasi_ditemukan': rec["lok_ditemukan"],
        'koordinat_ditemukan_teks': koordinat_ditemukan if koordinat_ditemukan != '-' else None,
        'tipe_ditemukan': temu_des["tipe"] if temu_des["tipe"] != '-' else None,
        'latitude_ditemukan': temu_des["lat"],
        'longitude_ditemukan': temu_des["lon"],
        'lainlain': rec["lain_lain"],
        'biaya_rp': fmt_num(rec["biaya"]) if rec["biaya"] is not None else None,
        'aksi': rec["aksi"],
        'durasi_operasi_hari_sumber': durasi_operasi_hari,  # dipakai preview, bukan dikirim ke normalize_payload
        # Metadata untuk pratinjau (bukan kolom kejadian_sar)
        '_meta': {'sheet': rec["sheet"], 'source_row': rec["source_row"]},
    }
    return payload


# ============================================================
# 8. Entry point: baca workbook -> daftar payload
# ============================================================

def parse_workbook(file_obj, tahun_default=None):
    """file_obj: file-like (mis. request.files['file'].stream) berisi .xlsx.
    tahun_default: dipakai untuk sheet yang nama-nya tidak mengandung tahun
    (mis. 'SEPTEMBER' tanpa '2023' di belakangnya).

    Return: list of dict, satu per kejadian, masing-masing dengan key
    'payload' (siap dilempar ke normalize_payload, setelah field '_meta' &
    'durasi_operasi_hari_sumber' dibuang) dan 'meta' (info sheet/baris asal
    untuk pesan error yang jelas ke admin)."""
    # read_only=True kadang membuat ws.max_row/max_column jadi None untuk
    # file tertentu (tergantung cara file disimpan) -- pakai mode normal
    # supaya dimensi sheet selalu terbaca benar. File laporan bulanan/
    # tahunan ukurannya kecil, jadi tidak masalah dari sisi memori/waktu.
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    hasil = []
    for sheet_name in wb.sheetnames:
        upper_name = sheet_name.upper()
        if not any(m in upper_name for m in MONTHS):
            continue  # lewati sheet non-bulan (mis. rekap "Per Unit")

        year_match = re.search(r'(20\d{2})', sheet_name)
        tahun_sheet = int(year_match.group(1)) if year_match else tahun_default
        if not tahun_sheet:
            # Tidak ada tahun di nama sheet & tidak ada default dari form --
            # tetap parse, tapi field waktu tidak bisa distandardisasi penuh
            # (tanggal MMDD tanpa info tahun jadi ambigu).
            tahun_sheet = datetime.now().year

        ws = wb[sheet_name]
        records = parse_month_sheet(ws, tahun_sheet)
        for rec in records:
            payload = build_payload_row(rec)
            meta = payload.pop('_meta')
            # 'durasi_operasi_hari_sumber' SENGAJA dibiarkan ada di payload --
            # ini nama field yang dibaca langsung oleh normalize_payload() di
            # routes/admin_routes.py untuk memutuskan sumber durasi_operasi_hari.
            hasil.append({'payload': payload, 'meta': meta})
    wb.close()
    return hasil
